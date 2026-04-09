import io, os
from datetime import datetime, timedelta, date
import numpy as np
import mysql.connector
from flask import (Flask, render_template, request, redirect,
                   url_for, flash, session, send_file, jsonify)
from flask_mail import Mail, Message
try:
    from flask_socketio import SocketIO, emit, join_room
    HAS_SOCKETIO = True
except ImportError:
    HAS_SOCKETIO = False
from sklearn.linear_model import LinearRegression
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (HRFlowable, Paragraph, SimpleDocTemplate,
                                 Spacer, Table, TableStyle)

app = Flask(__name__)
app.secret_key = "super_secret_blood_bank_key"
socketio = SocketIO(app, cors_allowed_origins="*", async_mode="threading") if HAS_SOCKETIO else None

db_config = {"host":"localhost","user":"root","password":"12345","database":"blood_bank_db"}

app.config.update(
    MAIL_SERVER="smtp.gmail.com", MAIL_PORT=587, MAIL_USE_TLS=True,
    MAIL_USERNAME=os.environ.get("MAIL_USERNAME","your_email@gmail.com"),
    MAIL_PASSWORD=os.environ.get("MAIL_PASSWORD","your_app_password"),
    MAIL_DEFAULT_SENDER=("LifeFlow", os.environ.get("MAIL_USERNAME","your_email@gmail.com")),
)
mail = Mail(app)

LOW_STOCK_THRESHOLD = 500
EXPIRY_DAYS = 35
ELIGIBILITY_DAYS = 56

BADGES = {
    "first_drop": {"label":"First Drop","icon":"🩸","desc":"Made your first donation"},
    "triple":     {"label":"Triple Saver","icon":"🏅","desc":"Donated 3 times"},
    "hero":       {"label":"Life Hero","icon":"🦸","desc":"Donated 5 times"},
    "legend":     {"label":"Blood Legend","icon":"🏆","desc":"Donated 10 times"},
    "rare_donor": {"label":"Rare Donor","icon":"💎","desc":"Rare blood group donor"},
}

HEALTH_TIPS = [
    {"tip":"Drink at least 500ml of water before donating blood.","icon":"💧"},
    {"tip":"Eat iron-rich foods like spinach and lentils to replenish after donation.","icon":"🥗"},
    {"tip":"Avoid alcohol for 24 hours before and after donation.","icon":"🚫"},
    {"tip":"Get a good night sleep before your donation day.","icon":"😴"},
    {"tip":"Rest for 10-15 minutes after donating.","icon":"🛋"},
    {"tip":"One blood donation can save up to 3 lives!","icon":"❤️"},
]

COMPATIBILITY = {
    "O-": {"donate_to":["A+","A-","B+","B-","AB+","AB-","O+","O-"],"receive_from":["O-"]},
    "O+": {"donate_to":["A+","B+","AB+","O+"],"receive_from":["O+","O-"]},
    "A-": {"donate_to":["A+","A-","AB+","AB-"],"receive_from":["A-","O-"]},
    "A+": {"donate_to":["A+","AB+"],"receive_from":["A+","A-","O+","O-"]},
    "B-": {"donate_to":["B+","B-","AB+","AB-"],"receive_from":["B-","O-"]},
    "B+": {"donate_to":["B+","AB+"],"receive_from":["B+","B-","O+","O-"]},
    "AB-":{"donate_to":["AB+","AB-"],"receive_from":["A-","B-","AB-","O-"]},
    "AB+":{"donate_to":["AB+"],"receive_from":["A+","A-","B+","B-","AB+","AB-","O+","O-"]},
}

BLOOD_BANK_LAT  = 11.3410
BLOOD_BANK_LNG  = 77.7172
BLOOD_BANK_NAME = "LifeFlow Central Blood Bank, Erode"
AVG_SPEED_KMPH  = 40

GEMINI_API_KEY  = "your-gemini-key-here"   # ← Replace with your Google Gemini Key

CHATBOT_KB = [
    {"keys":["eligible","when can i donate","how long","wait","56","days","next donation"],
     "reply":"You must wait **56 days** (8 weeks) between whole blood donations. Check your Eligibility Tracker above — it shows your exact countdown timer!"},
    {"keys":["first time","never donated","new donor","first donation"],
     "reply":"Welcome! If this is your first donation, you are fully eligible today. Make sure you are at least 18 years old, weigh over 50 kg, and are in good health. Log your donation using the form on this page."},
    {"keys":["o negative","o-","universal donor","universal"],
     "reply":"**O-** is the universal donor — it can be given to any blood group in emergencies. It is the most in-demand and rarest type. If you are O-, your donations are especially valuable!"},
    {"keys":["ab positive","ab+","universal recipient"],
     "reply":"**AB+** is the universal recipient — people with AB+ can receive blood from any blood group. AB+ donors can donate to AB+ recipients only."},
    {"keys":["blood group","blood type","my type","what is","compatible","compatibility"],
     "reply":"Use the **Blood Compatibility Checker** on this page! Select your blood group to instantly see which groups you can donate to and receive from."},
    {"keys":["how much","500","volume","ml","quantity","amount"],
     "reply":"Each whole blood donation is **500 ml** — roughly 10% of the average adult's blood volume. Your body replenishes this within 24–48 hours. The red blood cells are fully replaced within 56 days."},
    {"keys":["does it hurt","pain","needle","scared","nervous","afraid"],
     "reply":"Most donors feel only a small pinch from the needle, lasting 1–2 seconds. The actual donation takes about 8–10 minutes. The entire process from registration to refreshments is usually 45–60 minutes. Thousands of donors do it regularly!"},
    {"keys":["how long does it take","duration","time","minutes","process"],
     "reply":"The actual blood draw takes **8–10 minutes**. With registration and the brief health check, the whole visit is about **45–60 minutes**. Bring a book or music!"},
    {"keys":["eat","food","meal","before","diet","hungry"],
     "reply":"**Before donating:** Eat a healthy iron-rich meal at least 2 hours before. Avoid fatty foods. Drink at least 500ml of water beforehand. Do not donate on an empty stomach."},
    {"keys":["drink","water","hydrate","hydration","juice"],
     "reply":"**Stay hydrated!** Drink at least 500ml of water or juice before donating. Avoid alcohol for 24 hours before and after. Good hydration makes the donation faster and reduces dizziness."},
    {"keys":["after","rest","recovery","tired","dizzy","faint","feel"],
     "reply":"**After donating:** Rest for 10–15 minutes, enjoy the refreshments provided. Drink extra fluids for the next 24 hours. Avoid heavy lifting or strenuous exercise for the rest of the day. If you feel dizzy, sit or lie down immediately."},
    {"keys":["certificate","certif","download","proof"],
     "reply":"You can download your **Donation Certificate** (PDF) using the button on this page! It shows your name, blood group, donation date, and total donations. Perfect for records or recognition."},
    {"keys":["badge","achievement","reward","points","hero","legend"],
     "reply":"LifeFlow awards badges for milestones: 🩸 First Drop (1 donation), 🏅 Triple Saver (3), 🦸 Life Hero (5), 🏆 Blood Legend (10), 💎 Rare Donor (rare blood group). Check your Achievements section above!"},
    {"keys":["hello","hi","hey","good morning","good afternoon","good evening","help"],
     "reply":"Hello! 👋 I am the LifeFlow Donor Assistant. I can answer questions about donation eligibility, blood types, what to eat before donating, certificates, badges, and more. What would you like to know?"},
]

def get_db():
    return mysql.connector.connect(**db_config)

def log_activity(admin_id, action, details=""):
    try:
        conn=get_db(); cursor=conn.cursor()
        cursor.execute("INSERT INTO Admin_Activity_Log (admin_id,action,details) VALUES (%s,%s,%s)",
                       (admin_id,action,details))
        conn.commit(); cursor.close(); conn.close()
    except Exception as e:
        print(f"[Log] {e}")

def award_badges(user_id, total_donations, blood_group):
    earned=[]
    rules=[("first_drop",total_donations>=1),("triple",total_donations>=3),
           ("hero",total_donations>=5),("legend",total_donations>=10),
           ("rare_donor",blood_group in ["AB-","B-","A-","O-"])]
    try:
        conn=get_db(); cursor=conn.cursor()
        for key,cond in rules:
            if cond:
                cursor.execute("INSERT IGNORE INTO Donor_Achievements (user_id,badge_key) VALUES (%s,%s)",(user_id,key))
                if cursor.rowcount>0: earned.append(BADGES[key]["label"])
        conn.commit(); cursor.close(); conn.close()
    except Exception as e: print(f"[Badge] {e}")
    return earned

def check_low_stock_and_alert():
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT blood_group,quantity_ml FROM Blood_Inventory WHERE quantity_ml<%s",(LOW_STOCK_THRESHOLD,))
        low=cursor.fetchall()
        cursor.execute("SELECT email FROM Users WHERE role='Admin'")
        admins=[r["email"] for r in cursor.fetchall()]
        cursor.close(); conn.close()
        if not low or not admins: return
        rows="".join(f"<tr><td style='padding:6px 10px;'><b style='color:#C0152A;'>{r['blood_group']}</b></td><td style='padding:6px 10px;color:#F0970A;'>{r['quantity_ml']} ml</td></tr>" for r in low)
        html=f"<div style='font-family:sans-serif;'><h2>⚠ Low Stock Alert</h2><table>{rows}</table></div>"
        msg=Message(subject="⚠ LifeFlow: Low Blood Stock Alert",recipients=admins,html=html)
        mail.send(msg)
    except Exception as e: print(f"[LowStock] {e}")

def run_ml_demand_prediction():
    predictions=[]
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("""SELECT DATE(request_date) AS req_day, blood_group,
                          SUM(quantity_needed_ml) AS daily_ml
                          FROM Blood_Requests WHERE request_date>=DATE_SUB(CURDATE(),INTERVAL 30 DAY)
                          GROUP BY DATE(request_date),blood_group ORDER BY req_day ASC""")
        rows=cursor.fetchall()
        cursor.execute("SELECT blood_group,quantity_ml FROM Blood_Inventory")
        inventory={r["blood_group"]:r["quantity_ml"] for r in cursor.fetchall()}
        cursor.close(); conn.close()
        if not rows: return []
        all_dates=sorted({r["req_day"] for r in rows})
        date_index={d:i for i,d in enumerate(all_dates)}
        series_map={bg:{} for bg in inventory}
        for r in rows:
            bg=r["blood_group"]
            if bg in series_map: series_map[bg][r["req_day"]]=float(r["daily_ml"])
        for bg,series in series_map.items():
            current_stock=inventory.get(bg,0)
            if len(series)<3:
                avg=sum(series.values())/len(series) if series else 0
                predicted_7d=avg*7; slope=0.0
            else:
                X=np.array([[date_index[d]] for d in sorted(series)])
                y=np.array([series[d] for d in sorted(series)])
                model=LinearRegression(); model.fit(X,y)
                last_idx=max(date_index.values())
                future_X=np.array([[last_idx+i] for i in range(1,8)])
                predicted_7d=float(np.sum(np.clip(model.predict(future_X),0,None)))
                slope=float(model.coef_[0])
            shortage=max(0,predicted_7d-current_stock)
            trend="rising" if slope>50 else ("falling" if slope<-50 else "stable")
            predictions.append({"blood_group":bg,"current_stock":current_stock,
                                 "predicted_demand":round(predicted_7d),
                                 "predicted_shortage":round(shortage),"trend":trend})
        predictions.sort(key=lambda p:p["predicted_shortage"],reverse=True)
    except Exception as e: print(f"[ML] {e}")
    return predictions

def send_emergency_donor_alert(blood_group, requester_name, volume_ml):
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("""SELECT u.name,u.email FROM Users u JOIN Donors d ON u.user_id=d.user_id
                          WHERE d.blood_group=%s AND u.role='Donor' AND u.email IS NOT NULL
                          AND (d.last_donation_date IS NULL OR d.last_donation_date<=DATE_SUB(CURDATE(),INTERVAL 56 DAY))""",
                       (blood_group,))
        donors=cursor.fetchall(); cursor.close(); conn.close()
        if not donors: return 0,None
        sent=0
        for donor in donors:
            html=f"""<div style='font-family:sans-serif;background:#0D0D0F;color:#F0F0F4;max-width:520px;margin:auto;border:1px solid rgba(192,21,42,.3);border-radius:14px;overflow:hidden;'>
              <div style='background:#8B0E1E;padding:24px;text-align:center;'><h2 style='color:#fff;margin:0;'>🩸 Emergency Donor Alert</h2></div>
              <div style='padding:28px;'><p style='color:#8E8EA0;'>Dear <b style='color:#F0F0F4;'>{donor['name']}</b>,<br><br>
              <b>{requester_name}</b> urgently needs <b style='color:#C0152A;'>{blood_group}</b> blood ({volume_ml}ml). Your blood type matches!</p>
              <p style='color:#8E8EA0;font-size:13px;'>Please visit the blood bank at your earliest convenience.</p></div></div>"""
            msg=Message(subject=f"🚨 URGENT: {blood_group} Blood Needed — LifeFlow",recipients=[donor["email"]],html=html)
            mail.send(msg); sent+=1
        return sent,None
    except Exception as e: print(f"[Email] {e}"); return 0,str(e)

def generate_admin_pdf_report():
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=A4,leftMargin=2*cm,rightMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
    CR=colors.HexColor("#C0152A"); CR_DK=colors.HexColor("#8B0E1E")
    CARD=colors.HexColor("#16161B"); ELEV=colors.HexColor("#1D1D24")
    BDR=colors.HexColor("#2E2E3A"); TXT=colors.HexColor("#F0F0F4")
    MUT=colors.HexColor("#8E8EA0"); OK=colors.HexColor("#1BB86A")
    WRN=colors.HexColor("#F0970A"); ERR=colors.HexColor("#E8192F")
    def S(n,**kw):
        d=dict(fontName="Helvetica",fontSize=9,textColor=TXT,leading=13); d.update(kw)
        return ParagraphStyle(n,**d)
    s_title=S("t",fontName="Helvetica-Bold",fontSize=22,textColor=colors.white,alignment=TA_CENTER,leading=28)
    s_sub=S("sb",fontSize=10,textColor=MUT,alignment=TA_CENTER,leading=14)
    s_ts=S("ts",fontSize=8,textColor=colors.HexColor("#56566A"),alignment=TA_CENTER)
    s_h2=S("h2",fontName="Helvetica-Bold",fontSize=13,textColor=TXT,leading=18,spaceBefore=4)
    s_cell=S("c",fontSize=9,textColor=TXT,leading=12)
    s_csm=S("cs",fontSize=8,textColor=MUT,leading=11)
    s_cc=S("cc",fontSize=9,textColor=TXT,alignment=TA_CENTER,leading=12)
    def dyn(n,c,center=True):
        return S(n,fontSize=9,textColor=c,alignment=(TA_CENTER if center else TA_LEFT),leading=12)
    story=[]
    hdr=Table([[Paragraph("LifeFlow Blood Bank Report",s_title)]],colWidths=[doc.width])
    hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),CR_DK),("TOPPADDING",(0,0),(-1,-1),20),
                              ("BOTTOMPADDING",(0,0),(-1,-1),14),("LEFTPADDING",(0,0),(-1,-1),20),("RIGHTPADDING",(0,0),(-1,-1),20)]))
    story+=[hdr,Spacer(1,.2*cm),Paragraph("Admin Operational Report",s_sub),
            Paragraph(f"Generated: {datetime.now().strftime('%A, %d %B %Y at %H:%M')}",s_ts),Spacer(1,.4*cm)]
    def section(t,icon=""):
        story.append(HRFlowable(width="100%",thickness=0.8,color=BDR,spaceAfter=5))
        story.append(Paragraph(f"{icon}  {t}",s_h2)); story.append(Spacer(1,.2*cm))
    def tstyle():
        return [("BACKGROUND",(0,0),(-1,0),ELEV),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                ("FONTSIZE",(0,0),(-1,0),8),("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),9),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),[CARD,ELEV]),("GRID",(0,0),(-1,-1),.4,BDR),
                ("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),
                ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("TEXTCOLOR",(0,0),(-1,-1),TXT)]
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM Blood_Inventory ORDER BY blood_group"); inventory=cursor.fetchall()
        cursor.execute("""SELECT r.request_id,u.name AS requester_name,r.blood_group,r.quantity_needed_ml,r.priority,r.request_date
                          FROM Blood_Requests r JOIN Users u ON r.hospital_user_id=u.user_id
                          WHERE r.status='Approved' AND r.request_date>=DATE_SUB(CURDATE(),INTERVAL 7 DAY)
                          ORDER BY r.request_date DESC"""); approved=cursor.fetchall()
        cursor.execute("SELECT name AS hospital_name,email,contact_number FROM Users WHERE role='Hospital' ORDER BY name"); hospitals=cursor.fetchall()
        cursor.close(); conn.close()
    except Exception as e: print(f"[PDF] {e}"); inventory=[]; approved=[]; hospitals=[]
    ml=run_ml_demand_prediction()
    section("Live Blood Inventory","🩸")
    total_ml=sum(r["quantity_ml"] for r in inventory)
    story.append(Paragraph(f"Total: <b>{total_ml:,} ml</b> ({total_ml/1000:.1f}L)",s_csm))
    story.append(Spacer(1,.2*cm))
    if inventory:
        hdr2=[Paragraph(h,s_csm) for h in ["Blood Group","Stock (ml)","Status"]]
        rows=[hdr2]
        for item in inventory:
            ml_v=item["quantity_ml"]
            if ml_v<500: st=Paragraph("⚠ Low",dyn("e",ERR))
            elif ml_v>3000: st=Paragraph("✓ Good",dyn("g",OK))
            else: st=Paragraph("~ OK",dyn("w",WRN))
            rows.append([Paragraph(f"<b>{item['blood_group']}</b>",dyn("bg",CR)),Paragraph(f"{ml_v:,}",s_cc),st])
        cw=[3*cm,4*cm,None]; cw[-1]=doc.width-sum(cw[:-1])
        t=Table(rows,colWidths=cw,repeatRows=1); t.setStyle(TableStyle(tstyle())); story.append(t)
    story+=[Spacer(1,.8*cm),HRFlowable(width="100%",thickness=.4,color=BDR,spaceAfter=6),
            Paragraph(f"LifeFlow | {datetime.now().strftime('%d %b %Y %H:%M')} | CONFIDENTIAL",s_ts)]
    doc.build(story); buf.seek(0); return buf

def generate_donor_certificate(donor_name, blood_group, donation_date, total_donations):
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=landscape(A4),leftMargin=2*cm,rightMargin=2*cm,topMargin=2*cm,bottomMargin=2*cm)
    CR=colors.HexColor("#C0152A"); GOLD=colors.HexColor("#D4A437")
    TXT_D=colors.HexColor("#1A1A2E"); MUT=colors.HexColor("#555577")
    def S(n,**kw):
        d=dict(fontName="Helvetica",fontSize=12,textColor=TXT_D,leading=16,alignment=TA_CENTER); d.update(kw)
        return ParagraphStyle(n,**d)
    story=[]
    story.append(Spacer(1,.5*cm))
    story.append(Paragraph("🩸  LifeFlow Blood Bank",S("brand",fontName="Helvetica-Bold",fontSize=18,textColor=CR)))
    story.append(Paragraph("CERTIFICATE OF APPRECIATION",S("head",fontName="Helvetica-Bold",fontSize=28,textColor=GOLD,leading=34,spaceBefore=10)))
    story.append(Paragraph("This certificate is proudly presented to",S("pre",fontSize=13,textColor=MUT,spaceBefore=6)))
    story.append(Paragraph(donor_name,S("name",fontName="Helvetica-Bold",fontSize=36,textColor=CR,leading=44,spaceBefore=6)))
    story.append(Paragraph(f"For their generous blood donation of <b>500 ml</b> of <b>{blood_group}</b> blood on <b>{donation_date}</b>.",
                           S("body",fontSize=12,textColor=TXT_D,spaceBefore=10,leading=18)))
    story.append(Paragraph(f"Total Donations: <b>{total_donations}</b>  ·  Blood Group: <b>{blood_group}</b>",
                           S("stats",fontSize=11,textColor=MUT,spaceBefore=8)))
    story.append(Spacer(1,.6*cm))
    story.append(HRFlowable(width=12*cm,thickness=1,color=GOLD,spaceAfter=10))
    story.append(Paragraph("LifeFlow Blood Bank Authority",S("sig",fontName="Helvetica-Bold",fontSize=11)))
    story.append(Paragraph(f"Issued: {datetime.now().strftime('%d %B %Y')}",S("date",fontSize=9,textColor=MUT,spaceBefore=4)))
    border=Table([[story]],colWidths=[doc.width])
    border.setStyle(TableStyle([("BOX",(0,0),(-1,-1),3,GOLD),("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#FEFEF8")),
                                 ("TOPPADDING",(0,0),(-1,-1),24),("BOTTOMPADDING",(0,0),(-1,-1),24)]))
    doc.build([border]); buf.seek(0); return buf

@app.route("/")
def index(): return redirect(url_for("login"))

@app.route("/login",methods=["GET","POST"])
def login():
    if request.method=="POST":
        email=request.form["email"]; pwd=request.form["password"]; role=request.form["role"]
        try:
            conn=get_db(); cursor=conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM Users WHERE email=%s AND password_hash=%s AND role=%s",(email,pwd,role))
            user=cursor.fetchone(); cursor.close(); conn.close()
            if user:
                session["user_id"]=user["user_id"]; session["name"]=user["name"]; session["role"]=user["role"]; session["email"]=user.get("email","")
                if role=="Admin": return redirect(url_for("admin_dashboard"))
                if role=="Donor": return redirect(url_for("donor_dashboard"))
                if role=="Hospital": return redirect(url_for("hospital_dashboard"))
            else: return render_template("login.html",error="Invalid Email, Password, or Role.")
        except mysql.connector.Error: return render_template("login.html",error="Database Connection Error.")
    return render_template("login.html")

@app.route("/register",methods=["GET","POST"])
def register():
    if request.method=="POST":
        name=request.form["name"]; email=request.form["email"]; pwd=request.form["password"]
        contact=request.form["contact"]; bg=request.form["blood_group"]
        age=request.form["age"]; weight=request.form["weight"]
        try:
            conn=get_db(); cursor=conn.cursor()
            cursor.execute("INSERT INTO Users (name,email,password_hash,role,contact_number) VALUES (%s,%s,%s,'Donor',%s)",(name,email,pwd,contact))
            uid=cursor.lastrowid
            cursor.execute("INSERT INTO Donors (user_id,blood_group,age,weight) VALUES (%s,%s,%s,%s)",(uid,bg,age,weight))
            conn.commit(); cursor.close(); conn.close()
            flash("Registration successful! Please log in.","success"); return redirect(url_for("login"))
        except mysql.connector.IntegrityError: return render_template("register.html",error="Email already registered.")
        except mysql.connector.Error: return render_template("register.html",error="A database error occurred.")
    return render_template("register.html")

@app.route("/register_hospital",methods=["POST"])
def register_hospital():
    hospital_name=request.form["hospital_name"]; email=request.form["email"]
    pwd=request.form["password"]; contact=request.form["contact"]
    try:
        conn=get_db(); cursor=conn.cursor()
        cursor.execute("INSERT INTO Users (name,email,password_hash,role,contact_number) VALUES (%s,%s,%s,'Hospital',%s)",(hospital_name,email,pwd,contact))
        conn.commit(); cursor.close(); conn.close()
        flash(f"Hospital registered! Please log in.","success"); return redirect(url_for("login"))
    except mysql.connector.IntegrityError: return render_template("register.html",error="Email already registered.")
    except mysql.connector.Error: return render_template("register.html",error="A database error occurred.")

@app.route("/logout")
def logout(): session.clear(); return redirect(url_for("login"))

@app.route("/delete_account", methods=["POST"])
def delete_account():
    if not session.get("user_id"):
        return jsonify({"success": False, "error": "Not logged in"})
    uid  = session["user_id"]
    role = session.get("role", "")
    try:
        conn = get_db(); cur = conn.cursor()
        if role == "Admin":
            cur.execute("DELETE FROM Users WHERE user_id=%s", (uid,))
        elif role == "Hospital":
            cur.execute("DELETE FROM Blood_Requests WHERE hospital_user_id=%s", (uid,))
            cur.execute("DELETE FROM Users WHERE user_id=%s", (uid,))
        else:  # Donor
            cur.execute("DELETE FROM Donor_Achievements WHERE user_id=%s", (uid,))
            cur.execute("DELETE FROM Donors WHERE user_id=%s", (uid,))
            cur.execute("DELETE FROM Blood_Requests WHERE hospital_user_id=%s AND requester_role='Donor'", (uid,))
            cur.execute("DELETE FROM Users WHERE user_id=%s", (uid,))
        conn.commit(); cur.close(); conn.close()
        session.clear()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/change_password",methods=["POST"])
def change_password():
    if "user_id" not in session: return redirect(url_for("login"))
    current=request.form["current_password"]; new_pwd=request.form["new_password"]; confirm=request.form["confirm_password"]
    if new_pwd!=confirm: flash("New passwords do not match.","error"); return redirect(request.referrer or url_for("login"))
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT user_id FROM Users WHERE user_id=%s AND password_hash=%s",(session["user_id"],current))
        user=cursor.fetchone()
        if not user: flash("Current password is incorrect.","error")
        else:
            cursor.execute("UPDATE Users SET password_hash=%s WHERE user_id=%s",(new_pwd,session["user_id"]))
            conn.commit(); flash("Password updated successfully!","success")
            if session["role"]=="Admin": log_activity(session["user_id"],"Password Changed",f"{session['name']} changed password")
        cursor.close(); conn.close()
    except mysql.connector.Error: flash("Database error.","error")
    return redirect(request.referrer or url_for("login"))

@app.route("/admin_dashboard")
def admin_dashboard():
    if "user_id" not in session or session["role"]!="Admin": return redirect(url_for("login"))
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM Blood_Inventory"); inventory_data=cursor.fetchall()
        cursor.execute("""SELECT r.request_id,u.name AS requester_name,r.requester_role,r.blood_group,
                          r.quantity_needed_ml,r.priority,r.request_date
                          FROM Blood_Requests r JOIN Users u ON r.hospital_user_id=u.user_id
                          WHERE r.status='Pending' AND r.requester_role='Hospital'
                          ORDER BY FIELD(r.priority,'Critical','Urgent','Normal'),r.request_date ASC""")
        pending_requests=cursor.fetchall()
        cursor.execute("SELECT user_id,name AS hospital_name,email,contact_number FROM Users WHERE role='Hospital' ORDER BY name")
        hospitals_data=cursor.fetchall()
        try:
            cursor.execute("SELECT * FROM Blood_Batches WHERE expiry_date>=CURDATE() ORDER BY expiry_date ASC")
            batches=cursor.fetchall()
        except Exception: batches=[]
        cursor.execute("""SELECT l.action,l.details,l.logged_at,u.name AS admin_name
                          FROM Admin_Activity_Log l JOIN Users u ON l.admin_id=u.user_id
                          ORDER BY l.logged_at DESC LIMIT 20""")
        activity_log=cursor.fetchall()
        cursor.execute("""SELECT DATE(request_date) AS day,COUNT(*) AS total,
                          SUM(CASE WHEN status='Approved' THEN 1 ELSE 0 END) AS approved
                          FROM Blood_Requests WHERE request_date>=DATE_SUB(CURDATE(),INTERVAL 7 DAY)
                          GROUP BY DATE(request_date) ORDER BY day ASC""")
        weekly_requests=cursor.fetchall()
        cursor.execute("""SELECT blood_group,SUM(quantity_needed_ml) AS total_demand
                          FROM Blood_Requests WHERE request_date>=DATE_SUB(CURDATE(),INTERVAL 30 DAY)
                          GROUP BY blood_group ORDER BY total_demand DESC""")
        demand_by_group=cursor.fetchall()
        cursor.close(); conn.close()
        ml_predictions=run_ml_demand_prediction()
        forecast=(ml_predictions[0] if ml_predictions and ml_predictions[0]["predicted_shortage"]>0 else None)
        conn2=get_db(); cursor2=conn2.cursor(dictionary=True)
        try:
            cursor2.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE status='Approved' AND requester_role='Hospital'"); total_approved=cursor2.fetchone()["c"]
            cursor2.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE status='Pending'  AND requester_role='Hospital'");  total_pending =cursor2.fetchone()["c"]
            cursor2.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE requester_role='Hospital'");                          total_requests=cursor2.fetchone()["c"]
            cursor2.execute("SELECT SUM(quantity_ml) AS s FROM Blood_Inventory");               total_stock   =cursor2.fetchone()["s"] or 0
        except Exception:
            total_approved=total_pending=total_requests=0; total_stock=0
        finally:
            try: cursor2.close(); conn2.close()
            except Exception: pass
        return render_template("admin_dashboard.html",name=session["name"],
                               admin_email=session.get("email",""),
                               inventory=inventory_data,
                               requests=pending_requests,hospitals=hospitals_data,
                               batches=batches,activity_log=activity_log,forecast=forecast,
                               ml_predictions=ml_predictions,weekly_requests=weekly_requests,
                               demand_by_group=demand_by_group,today=date.today(),
                               total_approved=total_approved,total_pending=total_pending,
                               total_requests=total_requests,total_stock=total_stock)
    except mysql.connector.Error as err: return f"Database error: {err}"

@app.route("/update_stock",methods=["POST"])
def update_stock():
    if "user_id" not in session or session["role"]!="Admin": return redirect(url_for("login"))
    blood_group=request.form["blood_group"]; action=request.form["action"]; volume=int(request.form["volume"])
    try:
        conn=get_db(); cursor=conn.cursor()
        if action=="add":
            cursor.execute("UPDATE Blood_Inventory SET quantity_ml=quantity_ml+%s WHERE blood_group=%s",(volume,blood_group))
            expiry=date.today()+timedelta(days=EXPIRY_DAYS)
            try: cursor.execute("INSERT INTO Blood_Batches (blood_group,quantity_ml,collected_date,expiry_date) VALUES (%s,%s,CURDATE(),%s)",(blood_group,volume,expiry))
            except Exception: pass
            flash(f"Added {volume} ml to {blood_group}.","success")
            log_activity(session["user_id"],"Stock Added",f"{volume}ml of {blood_group}")
        elif action=="subtract":
            cursor.execute("UPDATE Blood_Inventory SET quantity_ml=GREATEST(0,quantity_ml-%s) WHERE blood_group=%s",(volume,blood_group))
            flash(f"Removed {volume} ml from {blood_group}.","success")
            log_activity(session["user_id"],"Stock Removed",f"{volume}ml of {blood_group}")
        conn.commit(); cursor.close(); conn.close(); check_low_stock_and_alert()
        if HAS_SOCKETIO and socketio:
            socketio.emit("stats_update", get_admin_stats(), room="admins")
    except mysql.connector.Error: flash("Error updating stock.","error")
    return redirect(url_for("admin_dashboard"))

@app.route("/process_request",methods=["POST"])
def process_request():
    if "user_id" not in session or session["role"]!="Admin": return redirect(url_for("login"))
    req_id=request.form["request_id"]; action=request.form["action"]
    bg=request.form.get("blood_group"); volume=int(request.form.get("volume",0))
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT hospital_user_id FROM Blood_Requests WHERE request_id=%s",(req_id,))
        req_row=cursor.fetchone()
        hospital_user_id=req_row["hospital_user_id"] if req_row else None
        if action=="approve":
            cursor.execute("SELECT quantity_ml FROM Blood_Inventory WHERE blood_group=%s",(bg,))
            stock=cursor.fetchone()
            if stock and stock["quantity_ml"]>=volume:
                cursor.execute("UPDATE Blood_Inventory SET quantity_ml=quantity_ml-%s WHERE blood_group=%s",(volume,bg))
                cursor.execute("UPDATE Blood_Requests SET status='Approved',processed_by=%s,processed_at=NOW() WHERE request_id=%s",(session["user_id"],req_id))
                flash(f"Approved request #{req_id}.","success")
                log_activity(session["user_id"],"Request Approved",f"#{req_id} {volume}ml {bg}")
                check_low_stock_and_alert()
                try:
                    h_id = hospital_user_id
                    cursor.execute("SELECT name FROM Users WHERE user_id=%s",(h_id,))
                    h_row = cursor.fetchone()
                    h_name = h_row["name"] if h_row else "Hospital"
                    eta = estimate_eta(h_name)
                    if HAS_SOCKETIO and socketio:
                        socketio.emit("request_approved",{
                            "request_id":  req_id,"blood_group": bg,"volume": volume,
                            "eta_text":    eta["eta_text"],"distance_km": eta["distance_km"],
                            "admin_name":  session["name"],
                        }, room=f"hospital_{h_id}")
                        socketio.emit("request_processed",{"request_id":req_id,"action":"approved","blood_group":bg}, room="admins")
                        socketio.emit("stats_update", get_admin_stats(), room="admins")
                        socketio.emit("hospital_stats_update", get_hospital_stats(h_id), room=f"hospital_{h_id}")
                except Exception as eta_err: print(f"[ETA] {eta_err}")
            else: flash("Not enough stock!","error")
        elif action=="reject":
            cursor.execute("UPDATE Blood_Requests SET status='Rejected',processed_by=%s,processed_at=NOW() WHERE request_id=%s",(session["user_id"],req_id))
            flash(f"Rejected request #{req_id}.","success")
            log_activity(session["user_id"],"Request Rejected",f"#{req_id}")
            if hospital_user_id and HAS_SOCKETIO and socketio:
                socketio.emit("request_status_update",{
                    "request_id":req_id,"status":"Rejected","blood_group":bg,
                    "message":f"❌ Your request #{req_id} for {bg} has been REJECTED.",
                    "admin":session["name"],
                }, room=f"hospital_{hospital_user_id}")
                socketio.emit("stats_update", get_admin_stats(), room="admins")
                socketio.emit("hospital_stats_update", get_hospital_stats(hospital_user_id), room=f"hospital_{hospital_user_id}")
        conn.commit(); cursor.close(); conn.close()
    except mysql.connector.Error: flash("Database error.","error")
    return redirect(url_for("admin_dashboard"))

@app.route("/download_report")
def download_report():
    if "user_id" not in session or session["role"]!="Admin": return redirect(url_for("login"))
    try:
        buf=generate_admin_pdf_report(); fn=f"LifeFlow_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        log_activity(session["user_id"],"PDF Downloaded",f"by {session['name']}")
        return send_file(buf,as_attachment=True,download_name=fn,mimetype="application/pdf")
    except Exception as e: print(f"[PDF] {e}"); flash("Could not generate report.","error"); return redirect(url_for("admin_dashboard"))

@app.route("/export_inventory_excel")
def export_inventory_excel():
    if "user_id" not in session or session["role"]!="Admin": return redirect(url_for("login"))
    try:
        import openpyxl
        from openpyxl.styles import Font,PatternFill,Alignment
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM Blood_Inventory ORDER BY blood_group"); inv=cursor.fetchall()
        cursor.close(); conn.close()
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Blood Inventory"
        red=PatternFill("solid",fgColor="C0152A"); hdr_font=Font(bold=True,color="FFFFFF")
        for i,h in enumerate(["Blood Group","Stock (ml)","Stock (L)","Status"],1):
            c=ws.cell(1,i,h); c.fill=red; c.font=hdr_font; c.alignment=Alignment(horizontal="center")
        for ri,item in enumerate(inv,2):
            ml=item["quantity_ml"]; status="Low" if ml<500 else ("Good" if ml>3000 else "OK")
            ws.cell(ri,1,item["blood_group"]); ws.cell(ri,2,ml); ws.cell(ri,3,round(ml/1000,2)); ws.cell(ri,4,status)
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        log_activity(session["user_id"],"Excel Export",f"by {session['name']}")
        return send_file(buf,as_attachment=True,download_name=f"LifeFlow_Inventory_{date.today()}.xlsx",
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e: print(f"[Excel] {e}"); flash("Install openpyxl first: pip install openpyxl","error"); return redirect(url_for("admin_dashboard"))

@app.route("/api/analytics")
def api_analytics():
    if "user_id" not in session or session["role"]!="Admin": return jsonify({}),403
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("""SELECT DATE(request_date) AS day,COUNT(*) AS total,
                          SUM(CASE WHEN status='Approved' THEN 1 ELSE 0 END) AS approved,
                          SUM(CASE WHEN status='Rejected' THEN 1 ELSE 0 END) AS rejected
                          FROM Blood_Requests WHERE request_date>=DATE_SUB(CURDATE(),INTERVAL 30 DAY)
                          GROUP BY DATE(request_date) ORDER BY day ASC""")
        daily=cursor.fetchall()
        cursor.execute("""SELECT blood_group,SUM(quantity_needed_ml) AS demand FROM Blood_Requests
                          WHERE request_date>=DATE_SUB(CURDATE(),INTERVAL 30 DAY) GROUP BY blood_group""")
        by_group=cursor.fetchall()
        cursor.execute("SELECT blood_group,quantity_ml FROM Blood_Inventory"); inv=cursor.fetchall()
        cursor.close(); conn.close()
        return jsonify({"daily":[{"day":str(r["day"]),"total":r["total"],"approved":int(r["approved"] or 0),"rejected":int(r["rejected"] or 0)} for r in daily],
                        "by_group":[{"bg":r["blood_group"],"demand":int(r["demand"] or 0)} for r in by_group],
                        "inventory":[{"bg":r["blood_group"],"ml":r["quantity_ml"]} for r in inv]})
    except Exception as e: return jsonify({"error":str(e)}),500

@app.route("/donor_dashboard")
def donor_dashboard():
    if "user_id" not in session or session["role"]!="Donor": return redirect(url_for("login"))
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM Donors WHERE user_id=%s",(session["user_id"],)); donor_info=cursor.fetchone()
        today=date.today(); last_date=donor_info["last_donation_date"] if donor_info else None
        eligible_date=(last_date+timedelta(days=ELIGIBILITY_DAYS)) if last_date else None
        days_remaining=(eligible_date-today).days if eligible_date and eligible_date>today else 0
        is_eligible=(last_date is None) or (today>=eligible_date)
        progress=min(int(((today-last_date).days/ELIGIBILITY_DAYS)*100),100) if last_date else 100
        cursor.execute("SELECT badge_key FROM Donor_Achievements WHERE user_id=%s",(session["user_id"],))
        earned_keys={r["badge_key"] for r in cursor.fetchall()}
        badges=[{**BADGES[k],"key":k,"earned":k in earned_keys} for k in BADGES]
        cursor.execute("""SELECT u.name,d.blood_group,d.total_donations FROM Users u JOIN Donors d ON u.user_id=d.user_id
                          WHERE d.total_donations>0 ORDER BY d.total_donations DESC LIMIT 10""")
        leaderboard=cursor.fetchall()
        cursor.execute("SELECT * FROM Blood_Requests WHERE hospital_user_id=%s AND requester_role='Donor' ORDER BY request_date DESC",(session["user_id"],))
        my_blood_requests=cursor.fetchall()
        cursor.execute("SELECT blood_group,quantity_ml FROM Blood_Inventory")
        availability={r["blood_group"]:r["quantity_ml"] for r in cursor.fetchall()}
        cursor.execute("SELECT name AS hospital_name,email,contact_number FROM Users WHERE role='Hospital' ORDER BY name")
        nearby_hospitals=cursor.fetchall()
        cursor.close(); conn.close()
        return render_template("donor_dashboard.html",name=session["name"],donor_info=donor_info,
                               last_donation_date=last_date,eligible_date=eligible_date,
                               days_remaining=days_remaining,is_eligible=is_eligible,progress=progress,today=today,
                               badges=badges,earned_keys=earned_keys,leaderboard=leaderboard,
                               my_blood_requests=my_blood_requests,availability=availability,
                               nearby_hospitals=nearby_hospitals,health_tips=HEALTH_TIPS,compatibility=COMPATIBILITY)
    except mysql.connector.Error as err: return f"Database Error: {err}"

@app.route("/log_donation",methods=["POST"])
def log_donation():
    if "user_id" not in session or session["role"]!="Donor": return redirect(url_for("login"))
    bg=request.form["blood_group"]
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT total_donations,blood_group FROM Donors WHERE user_id=%s",(session["user_id"],))
        row=cursor.fetchone(); new_total=(row["total_donations"] or 0)+1 if row else 1
        cursor.execute("UPDATE Donors SET last_donation_date=CURRENT_DATE(),total_donations=%s WHERE user_id=%s",(new_total,session["user_id"]))
        cursor.execute("UPDATE Blood_Inventory SET quantity_ml=quantity_ml+500 WHERE blood_group=%s",(bg,))
        conn.commit()
        new_badges=award_badges(session["user_id"],new_total,bg)
        if new_badges: flash(f"🏅 Badge earned: {', '.join(new_badges)}!","success")
        cursor.close(); conn.close(); check_low_stock_and_alert()
        flash("Thank you! 500 ml added to inventory.","success")
        if HAS_SOCKETIO and socketio:
            donor_id=session["user_id"]
            all_badges_now=[]
            try:
                c2=get_db().cursor(dictionary=True)
                c2.execute("SELECT badge_key FROM Donor_Achievements WHERE user_id=%s",(donor_id,))
                all_badges_now=[r["badge_key"] for r in c2.fetchall()]; c2.close()
            except Exception: pass
            socketio.emit("donor_stats_update",{"total_donations":new_total,"badges":all_badges_now}, room=f"donor_{donor_id}")
            for b in new_badges:
                badge_key=[k for k,v in BADGES.items() if v["label"]==b]
                if badge_key:
                    bk=badge_key[0]
                    socketio.emit("badge_unlocked",{"key":bk,"label":BADGES[bk]["label"],"icon":BADGES[bk]["icon"],"desc":BADGES[bk]["desc"]}, room=f"donor_{donor_id}")
            socketio.emit("stats_update", get_admin_stats(), room="admins")
    except mysql.connector.Error: flash("Could not log donation.","error")
    return redirect(url_for("donor_dashboard"))

@app.route("/donor_request_blood",methods=["POST"])
def donor_request_blood():
    if "user_id" not in session or session["role"]!="Donor": return redirect(url_for("login"))
    bg=request.form["blood_group"]; volume=int(request.form["volume"])
    reason=request.form.get("reason","Personal medical need"); priority=request.form.get("priority","Normal")
    try:
        conn=get_db(); cursor=conn.cursor()
        cursor.execute("INSERT INTO Blood_Requests (hospital_user_id,requester_role,blood_group,quantity_needed_ml,priority,reason) VALUES (%s,'Donor',%s,%s,%s,%s)",
                       (session["user_id"],bg,volume,priority,reason))
        req_id=cursor.lastrowid
        conn.commit(); cursor.close(); conn.close()
        flash(f"Blood request for {volume}ml of {bg} submitted. Admin will review shortly.","success")
        if HAS_SOCKETIO and socketio:
            socketio.emit("stats_update", get_admin_stats(), room="admins")
            socketio.emit("hospital_stats_update", get_hospital_stats(session["user_id"]), room=f"hospital_{session['user_id']}")
    except mysql.connector.Error: flash("Could not submit request.","error")
    return redirect(url_for("donor_dashboard"))

@app.route("/donor_certificate")
def donor_certificate():
    if "user_id" not in session or session["role"]!="Donor": return redirect(url_for("login"))
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM Donors WHERE user_id=%s",(session["user_id"],)); d=cursor.fetchone()
        cursor.close(); conn.close()
        if not d or not d["last_donation_date"]:
            flash("You need at least one donation to download a certificate.","error"); return redirect(url_for("donor_dashboard"))
        buf=generate_donor_certificate(donor_name=session["name"],blood_group=d["blood_group"],
                                       donation_date=d["last_donation_date"].strftime("%d %B %Y"),
                                       total_donations=d["total_donations"] or 1)
        fn=f"LifeFlow_Certificate_{session['name'].replace(' ','_')}.pdf"
        return send_file(buf,as_attachment=True,download_name=fn,mimetype="application/pdf")
    except Exception as e: print(f"[Cert] {e}"); flash("Could not generate certificate.","error"); return redirect(url_for("donor_dashboard"))

@app.route("/api/health_tip")
def api_health_tip():
    import random; tip=random.choice(HEALTH_TIPS); return jsonify(tip)

@app.route("/api/compatibility/<blood_group>")
def api_compatibility(blood_group):
    return jsonify(COMPATIBILITY.get(blood_group,{}))

@app.route("/hospital_dashboard")
def hospital_dashboard():
    if "user_id" not in session or session["role"]!="Hospital": return redirect(url_for("login"))
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT name,email,contact_number FROM Users WHERE user_id=%s",(session["user_id"],))
        profile=cursor.fetchone()
        hospital_name=profile["name"] if profile else session["name"]
        hospital_email=profile["email"] if profile else ""
        hospital_contact=profile["contact_number"] if profile else ""
        session["name"]=hospital_name
        cursor.execute("""SELECT r.*,u.name AS approved_by_name FROM Blood_Requests r
                          LEFT JOIN Users u ON r.processed_by=u.user_id
                          WHERE r.hospital_user_id=%s AND r.requester_role='Hospital'
                          ORDER BY r.request_date DESC""",(session["user_id"],))
        hospital_requests=cursor.fetchall()
        cursor.execute("""SELECT r.request_id,u.name AS donor_name,r.blood_group,
                          r.quantity_needed_ml,r.priority,r.reason,r.request_date,r.status
                          FROM Blood_Requests r JOIN Users u ON r.hospital_user_id=u.user_id
                          WHERE r.requester_role='Donor'
                          ORDER BY FIELD(r.priority,'Critical','Urgent','Normal'),r.request_date DESC
                          LIMIT 20""")
        donor_requests=cursor.fetchall()
        cursor.execute("SELECT blood_group,quantity_ml FROM Blood_Inventory")
        availability={r["blood_group"]:r["quantity_ml"] for r in cursor.fetchall()}
        cursor.close(); conn.close()
        return render_template("hospital_dashboard.html",name=hospital_name,hospital_name=hospital_name,
                               hospital_email=hospital_email,hospital_contact=hospital_contact,
                               requests=hospital_requests,donor_requests=donor_requests,
                               availability=availability)
    except mysql.connector.Error as err: return f"Database Error: {err}"

@app.route("/request_blood",methods=["POST"])
def request_blood():
    if "user_id" not in session or session["role"]!="Hospital": return redirect(url_for("login"))
    bg=request.form["blood_group"]; volume=int(request.form["volume"])
    priority=request.form.get("priority","Normal"); h_id=session["user_id"]
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("INSERT INTO Blood_Requests (hospital_user_id,requester_role,blood_group,quantity_needed_ml,priority) VALUES (%s,'Hospital',%s,%s,%s)",(h_id,bg,volume,priority))
        req_id=cursor.lastrowid
        cursor.execute("SELECT name FROM Users WHERE user_id=%s",(h_id,))
        row=cursor.fetchone(); hospital_name=row["name"] if row else "A Hospital"
        conn.commit(); cursor.close(); conn.close()
        flash(f"Blood request for {volume}ml of {bg} ({priority}) submitted.","success")
        priority_icon={"Critical":"🔴","Urgent":"🟡","Normal":"🟢"}.get(priority,"🔵")
        if HAS_SOCKETIO and socketio:
            socketio.emit("new_blood_request",{
                "request_id":    req_id,
                "hospital":      hospital_name,
                "blood_group":   bg,
                "volume":        volume,
                "priority":      priority,
                "priority_icon": priority_icon,
                "time":          datetime.now().strftime("%H:%M"),
                "message":       f"New {priority} Request from {hospital_name} — {bg} ({volume}ml)",
            }, room="admins")
            socketio.emit("stats_update", get_admin_stats(), room="admins")
            socketio.emit("hospital_stats_update", get_hospital_stats(h_id), room=f"hospital_{h_id}")
        sent,err=send_emergency_donor_alert(bg,hospital_name,volume)
        if sent>0: flash(f"📧 Alert sent to {sent} eligible {bg} donor(s)!","success")
        elif err: print(f"[Email] {err}")
    except mysql.connector.Error: flash("Could not submit request.","error")
    return redirect(url_for("hospital_dashboard"))

@app.route("/api/blood_availability")
def api_blood_availability():
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT blood_group,quantity_ml FROM Blood_Inventory")
        data={r["blood_group"]:r["quantity_ml"] for r in cursor.fetchall()}
        cursor.close(); conn.close(); return jsonify(data)
    except Exception as e: return jsonify({"error":str(e)}),500


@app.route("/admin_delete_user", methods=["POST"])
def admin_delete_user():
    if "user_id" not in session or session["role"] != "Admin":
        return redirect(url_for("login"))
    # Support delete by user_id (hospitals) or by email (donors)
    uid = request.form.get("user_id")
    email = request.form.get("user_id_by_email")
    try:
        conn = get_db(); cur = conn.cursor(dictionary=True)
        if uid:
            cur.execute("SELECT user_id, role, name FROM Users WHERE user_id=%s", (uid,))
        else:
            cur.execute("SELECT user_id, role, name FROM Users WHERE email=%s", (email,))
        user = cur.fetchone()
        if not user:
            flash("User not found.", "error")
            cur.close(); conn.close()
            return redirect(url_for("admin_dashboard"))
        del_id = user["user_id"]; del_role = user["role"]; del_name = user["name"]
        if del_role == "Donor":
            cur.execute("DELETE FROM Donor_Achievements WHERE user_id=%s", (del_id,))
            cur.execute("DELETE FROM Donors WHERE user_id=%s", (del_id,))
            cur.execute("DELETE FROM Blood_Requests WHERE hospital_user_id=%s AND requester_role='Donor'", (del_id,))
        elif del_role == "Hospital":
            cur.execute("DELETE FROM Blood_Requests WHERE hospital_user_id=%s", (del_id,))
        cur.execute("DELETE FROM Users WHERE user_id=%s", (del_id,))
        conn.commit(); cur.close(); conn.close()
        log_activity(session["user_id"], "User Deleted", f"{del_role}: {del_name} (id={del_id})")
        flash(f"{del_role} '{del_name}' deleted successfully.", "success")
    except Exception as e:
        flash(f"Error deleting user: {str(e)}", "error")
    return redirect(url_for("admin_dashboard"))

@socketio.on("join_admin_room")
def handle_join_admin_room(data):
    join_room("admins"); emit("joined",{"msg":"Connected to admin room"})

@socketio.on("join_admin")
def handle_join_admin(data):
    join_room("admins"); emit("joined",{"msg":"Connected to admin room"})

@socketio.on("join_hospital_room")
def handle_join_hospital_room(data):
    hospital_id=data.get("hospital_id")
    if hospital_id: join_room(f"hospital_{hospital_id}"); emit("joined",{"msg":"Connected to hospital room"})

@socketio.on("join_hospital")
def handle_join_hospital(data):
    room=f"hospital_{data.get('hospital_id','')}"; join_room(room); emit("joined",{"msg":"Connected to hospital room"})

def notify_admins(event_type,payload):
    if HAS_SOCKETIO and socketio: socketio.emit(event_type,payload,room="admins")

def notify_hospital(hospital_id,event_type,payload):
    if HAS_SOCKETIO and socketio: socketio.emit(event_type,payload,room=f"hospital_{hospital_id}")

@socketio.on("join_donor")
def handle_join_donor(data):
    donor_id=data.get("donor_id")
    if donor_id: join_room(f"donor_{donor_id}"); emit("joined",{"msg":"Connected to donor room"})

def get_admin_stats():
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE requester_role='Hospital'"); total=cursor.fetchone()["c"]
        cursor.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE status='Approved' AND requester_role='Hospital'"); approved=cursor.fetchone()["c"]
        cursor.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE status='Pending'  AND requester_role='Hospital'"); pending=cursor.fetchone()["c"]
        cursor.execute("SELECT COALESCE(SUM(quantity_ml),0) AS s FROM Blood_Inventory"); stock=cursor.fetchone()["s"]
        cursor.close(); conn.close()
        return {"total_requests":total,"total_approved":approved,"total_pending":pending,"total_stock":int(stock)}
    except Exception: return {}

def get_hospital_stats(user_id):
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE hospital_user_id=%s AND requester_role='Hospital'",(user_id,)); total=cursor.fetchone()["c"]
        cursor.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE hospital_user_id=%s AND status='Approved'",(user_id,)); approved=cursor.fetchone()["c"]
        cursor.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE hospital_user_id=%s AND status='Pending'",(user_id,)); pending=cursor.fetchone()["c"]
        cursor.execute("SELECT COUNT(*) AS c FROM Blood_Requests WHERE hospital_user_id=%s AND status='Rejected'",(user_id,)); rejected=cursor.fetchone()["c"]
        cursor.close(); conn.close()
        return {"total":total,"approved":approved,"pending":pending,"rejected":rejected}
    except Exception: return {}

@app.route("/api/admin_stats")
def api_admin_stats():
    if "user_id" not in session or session["role"]!="Admin": return jsonify({}),403
    return jsonify(get_admin_stats())

@app.route("/api/hospital_stats")
def api_hospital_stats():
    if "user_id" not in session or session["role"]!="Hospital": return jsonify({}),403
    return jsonify(get_hospital_stats(session["user_id"]))

@app.route("/api/donor_stats")
def api_donor_stats():
    if "user_id" not in session or session["role"]!="Donor": return jsonify({}),403
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT total_donations,blood_group FROM Donors WHERE user_id=%s",(session["user_id"],))
        row=cursor.fetchone()
        cursor.execute("SELECT badge_key FROM Donor_Achievements WHERE user_id=%s",(session["user_id"],))
        badges=[r["badge_key"] for r in cursor.fetchall()]
        cursor.close(); conn.close()
        return jsonify({"total_donations":row["total_donations"] if row else 0,"badges":badges})
    except Exception: return jsonify({}),500

@app.route("/api/chatbot",methods=["POST"])
def chatbot():
    data=request.get_json(); message=(data.get("message","") or "").strip(); history=data.get("history",[])
    if not message: return jsonify({"reply":"Please type a question and I will do my best to help!"})
    key=GEMINI_API_KEY
    user_name=session.get('name','Guest'); blood_group="unknown"
    if 'user_id' in session and session.get('role')=='Donor':
        try:
            conn=get_db(); cursor=conn.cursor(dictionary=True)
            cursor.execute("SELECT blood_group FROM Donors WHERE user_id=%s",(session['user_id'],))
            donor=cursor.fetchone(); cursor.close(); conn.close()
            if donor and donor['blood_group']: blood_group=donor['blood_group']
        except Exception as e: print(f"[Bot DB] {e}")
    if not key or key=="your-gemini-key-here":
        msg_lower=message.lower(); best_score,best_reply=0,None
        for entry in CHATBOT_KB:
            score=sum(1 for kw in entry["keys"] if kw in msg_lower)
            if score>best_score: best_score,best_reply=score,entry["reply"]
        if not best_reply: best_reply="I am not sure about that. Try asking about eligibility, blood groups, food before donating, certificates, or badges!"
        return jsonify({"reply":best_reply,"time":datetime.now().strftime("%H:%M")})
    system_prompt=(f"You are LifeBot, an expert AI medical assistant for LifeFlow Blood Bank. "
        f"User: {user_name}, blood group: {blood_group}. "
        "App features: 1) Request Blood button on dashboard. 2) Log Donation adds 500ml. 3) Download Certificate after first donation. "
        "Answer eligibility (18-65yrs, 50kg+, 56-day wait), recovery, blood matching. Use <b>bold</b> and <br> for formatting. Be warm and direct.")
    try:
        import google.generativeai as genai
        genai.configure(api_key=key)
        model=genai.GenerativeModel('gemini-2.5-flash',system_instruction=system_prompt)
        gemini_history=[]
        for h in history[-8:]:
            role="user" if h.get("role")=="user" else "model"
            content=h.get("content","")
            if content: gemini_history.append({"role":role,"parts":[content]})
        chat=model.start_chat(history=gemini_history)
        response=chat.send_message(message); ai_reply=response.text
    except Exception as e:
        print(f"[Chatbot] {e}"); ai_reply="Sorry, trouble connecting. Please try again!"
    return jsonify({"reply":ai_reply,"time":datetime.now().strftime("%H:%M")})

@app.route("/api/delivery_eta",methods=["POST"])
def delivery_eta():
    import math
    data=request.get_json(); hosp_lat=float(data.get("lat",0)); hosp_lng=float(data.get("lng",0))
    if not hosp_lat or not hosp_lng: return jsonify({"error":"No coordinates provided"}),400
    R=6371; dlat=math.radians(hosp_lat-BLOOD_BANK_LAT); dlng=math.radians(hosp_lng-BLOOD_BANK_LNG)
    a=(math.sin(dlat/2)**2+math.cos(math.radians(BLOOD_BANK_LAT))*math.cos(math.radians(hosp_lat))*math.sin(dlng/2)**2)
    dist_km=round(R*2*math.atan2(math.sqrt(a),math.sqrt(1-a)),1); eta_minutes=round((dist_km/40)*60)+10
    try:
        import urllib.request,json as _json
        url=(f"http://router.project-osrm.org/route/v1/driving/{BLOOD_BANK_LNG},{BLOOD_BANK_LAT};{hosp_lng},{hosp_lat}?overview=false")
        with urllib.request.urlopen(url,timeout=3) as resp:
            osrm=_json.loads(resp.read())
            if osrm.get("routes"): eta_minutes=round(osrm["routes"][0]["duration"]/60)+10; dist_km=round(osrm["routes"][0]["distance"]/1000,1)
    except Exception: pass
    hours=eta_minutes//60; minutes=eta_minutes%60
    return jsonify({"distance_km":dist_km,"eta_minutes":eta_minutes,"eta_str":f"{hours}h {minutes}min" if hours>0 else f"{minutes} minutes"})

@app.route("/api/request_eta/<int:request_id>")
def request_eta(request_id):
    if "user_id" not in session: return jsonify({}),403
    try:
        conn=get_db(); cursor=conn.cursor(dictionary=True)
        cursor.execute("SELECT r.status,r.processed_at,r.blood_group,r.quantity_needed_ml FROM Blood_Requests r WHERE r.request_id=%s",(request_id,))
        req=cursor.fetchone(); cursor.close(); conn.close()
        if not req or req["status"]!="Approved": return jsonify({"status":req["status"] if req else "unknown"})
        return jsonify({"status":"Approved","blood_group":req["blood_group"],"volume":req["quantity_needed_ml"],"processed_at":str(req["processed_at"])})
    except Exception as e: return jsonify({"error":str(e)}),500

def estimate_eta(hospital_name):
    import math,urllib.request,urllib.parse,json as _json
    try:
        q=urllib.parse.quote_plus(hospital_name+", India")
        url=f"https://nominatim.openstreetmap.org/search?q={q}&format=json&limit=1"
        req=urllib.request.Request(url,headers={"User-Agent":"LifeFlowApp/1.0"})
        with urllib.request.urlopen(req,timeout=4) as r:
            geo=_json.loads(r.read())
        hosp_lat=float(geo[0]["lat"]) if geo else BLOOD_BANK_LAT+0.05
        hosp_lng=float(geo[0]["lon"]) if geo else BLOOD_BANK_LNG+0.05
    except Exception: hosp_lat=BLOOD_BANK_LAT+0.05; hosp_lng=BLOOD_BANK_LNG+0.05
    dist_km=5.0; eta_minutes=20
    try:
        osrm_url=(f"http://router.project-osrm.org/route/v1/driving/{BLOOD_BANK_LNG},{BLOOD_BANK_LAT};{hosp_lng},{hosp_lat}?overview=false")
        req2=urllib.request.Request(osrm_url,headers={"User-Agent":"LifeFlowApp/1.0"})
        with urllib.request.urlopen(req2,timeout=5) as r2:
            osrm=_json.loads(r2.read())
        dist_km=round(osrm["routes"][0]["distance"]/1000,1); eta_minutes=round(osrm["routes"][0]["duration"]/60)+10
    except Exception:
        R=6371; dlat=math.radians(hosp_lat-BLOOD_BANK_LAT); dlng=math.radians(hosp_lng-BLOOD_BANK_LNG)
        a=(math.sin(dlat/2)**2+math.cos(math.radians(BLOOD_BANK_LAT))*math.cos(math.radians(hosp_lat))*math.sin(dlng/2)**2)
        dist_km=round(6371*2*math.asin(math.sqrt(a))*1.3,1); eta_minutes=round((dist_km/40)*60)+10
    h=eta_minutes//60; m=eta_minutes%60
    return {"distance_km":dist_km,"eta_minutes":eta_minutes,"eta_text":f"{h}h {m}m" if h>0 else f"{eta_minutes} minutes"}

if __name__=="__main__":
    if HAS_SOCKETIO and socketio:
        socketio.run(app, debug=True)
    else:
        app.run(debug=True)
